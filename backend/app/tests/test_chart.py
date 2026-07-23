import io

import openpyxl
from fastapi.testclient import TestClient
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, ScatterChart

from app.core.config import get_settings
from app.main import app

client = TestClient(app)


def _upload_sheets(sheets: dict[str, list[list]]) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)

    response = client.post(
        "/api/v1/upload",
        files={
            "file": (
                "test.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201
    return response.json()["file_id"]


def _apply(file_id: str, request: dict, commit: bool = False):
    url = f"/api/v1/workbook/{file_id}/chart"
    if commit:
        url += "?commit=true"
    return client.post(url, json=request)


def _load_stored_workbook(tmp_path, file_id: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(tmp_path / f"{file_id}.xlsx")


def _title_text(chart) -> str | None:
    if chart.title is None:
        return None
    return chart.title.tx.rich.p[0].r[0].t


_SHEET = {"Data": [["Label", "Value"], ["A", 1], ["B", 2], ["C", 3]]}


def test_chart_bar_preview_only(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(
        file_id,
        {
            "sheet_name": "Data",
            "chart_type": "bar",
            "data_range": "B1:B4",
            "categories_range": "A2:A4",
            "anchor": "D2",
            "title": "My Chart",
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["new_file_id"] is None
    assert body["chart_type"] == "bar"
    assert body["anchor"] == "D2"

    original = _load_stored_workbook(tmp_path, file_id)
    assert len(original["Data"]._charts) == 0


def test_chart_bar_commit_embeds_chart(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(
        file_id,
        {
            "sheet_name": "Data",
            "chart_type": "bar",
            "data_range": "B1:B4",
            "categories_range": "A2:A4",
            "anchor": "D2",
            "title": "My Chart",
        },
        commit=True,
    )

    assert response.status_code == 200
    new_file_id = response.json()["new_file_id"]
    assert new_file_id != file_id

    workbook = _load_stored_workbook(tmp_path, new_file_id)
    charts = workbook["Data"]._charts
    assert len(charts) == 1
    chart = charts[0]
    assert isinstance(chart, BarChart)
    assert _title_text(chart) == "My Chart"
    assert chart.anchor._from.col == 3  # D -> 0-indexed col 3
    assert chart.anchor._from.row == 1  # row 2 -> 0-indexed row 1

    original = _load_stored_workbook(tmp_path, file_id)
    assert len(original["Data"]._charts) == 0


def test_chart_line(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(
        file_id,
        {"sheet_name": "Data", "chart_type": "line", "data_range": "B1:B4", "categories_range": "A2:A4"},
        commit=True,
    )

    assert response.status_code == 200
    workbook = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert isinstance(workbook["Data"]._charts[0], LineChart)


def test_chart_area(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(
        file_id,
        {"sheet_name": "Data", "chart_type": "area", "data_range": "B1:B4", "categories_range": "A2:A4"},
        commit=True,
    )

    assert response.status_code == 200
    workbook = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert isinstance(workbook["Data"]._charts[0], AreaChart)


def test_chart_pie(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(
        file_id,
        {"sheet_name": "Data", "chart_type": "pie", "data_range": "B1:B4", "categories_range": "A2:A4"},
        commit=True,
    )

    assert response.status_code == 200
    workbook = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert isinstance(workbook["Data"]._charts[0], PieChart)


def test_chart_pie_rejects_multi_column_data_range(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A", "B", "C"], [1, 2, 3], [4, 5, 6]]})

    response = _apply(file_id, {"sheet_name": "Data", "chart_type": "pie", "data_range": "A1:C3"})

    assert response.status_code == 400


def test_chart_scatter(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["X", "Y"], [1, 2], [2, 4], [3, 6]]})

    response = _apply(
        file_id,
        {"sheet_name": "Data", "chart_type": "scatter", "x_range": "A2:A4", "y_range": "B2:B4"},
        commit=True,
    )

    assert response.status_code == 200
    workbook = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    charts = workbook["Data"]._charts
    assert len(charts) == 1
    assert isinstance(charts[0], ScatterChart)


def test_chart_scatter_rejects_data_range(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["X", "Y"], [1, 2]]})

    response = _apply(file_id, {"sheet_name": "Data", "chart_type": "scatter", "data_range": "A1:B2"})

    assert response.status_code == 422


def test_chart_bar_requires_data_range(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(file_id, {"sheet_name": "Data", "chart_type": "bar"})

    assert response.status_code == 422


def test_chart_invalid_data_range_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(file_id, {"sheet_name": "Data", "chart_type": "bar", "data_range": "not-a-range"})

    assert response.status_code == 400


def test_chart_data_range_out_of_bounds_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(file_id, {"sheet_name": "Data", "chart_type": "bar", "data_range": "B1:B50"})

    assert response.status_code == 400


def test_chart_invalid_anchor_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(
        file_id, {"sheet_name": "Data", "chart_type": "bar", "data_range": "B1:B4", "anchor": "not-a-cell"}
    )

    assert response.status_code == 400


def test_chart_invalid_sheet_name_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(_SHEET)

    response = _apply(file_id, {"sheet_name": "DoesNotExist", "chart_type": "bar", "data_range": "B1:B4"})

    assert response.status_code == 400


def test_chart_missing_file_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    response = _apply("does-not-exist", {"sheet_name": "Data", "chart_type": "bar", "data_range": "B1:B4"})

    assert response.status_code == 404
