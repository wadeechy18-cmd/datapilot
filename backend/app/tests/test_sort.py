import io

import openpyxl
from fastapi.testclient import TestClient

from app.core.config import get_settings
from app.main import app

client = TestClient(app)


def _upload_sheets(sheets: dict[str, list[list]]) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)

    response = client.post(
        "/api/v1/upload",
        files={
            "file": (
                "test.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201
    return response.json()["file_id"]


def _sort(file_id: str, request: dict, commit: bool = False):
    url = f"/api/v1/workbook/{file_id}/sort"
    if commit:
        url += "?commit=true"
    return client.post(url, json=request)


def _load_stored_workbook(tmp_path, file_id: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(tmp_path / f"{file_id}.xlsx")


def test_sort_ascending_preview_only(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name", "Score"], ["Carol", 30], ["Alice", 10], ["Bob", 20]]})

    response = _sort(file_id, {"sheet_name": "Data", "column": "B", "ascending": True})

    assert response.status_code == 200
    body = response.json()
    assert body["new_file_id"] is None
    assert body["row_count"] == 3

    original = _load_stored_workbook(tmp_path, file_id)
    assert [c.value for c in original["Data"]["A"]] == ["Name", "Carol", "Alice", "Bob"]


def test_sort_ascending_commit(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name", "Score"], ["Carol", 30], ["Alice", 10], ["Bob", 20]]})

    response = _sort(file_id, {"sheet_name": "Data", "column": "B", "ascending": True}, commit=True)

    assert response.status_code == 200
    body = response.json()
    new_file_id = body["new_file_id"]
    assert new_file_id != file_id

    sorted_wb = _load_stored_workbook(tmp_path, new_file_id)
    assert [c.value for c in sorted_wb["Data"]["A"]] == ["Name", "Alice", "Bob", "Carol"]
    assert [c.value for c in sorted_wb["Data"]["B"]] == ["Score", 10, 20, 30]

    original = _load_stored_workbook(tmp_path, file_id)
    assert [c.value for c in original["Data"]["A"]] == ["Name", "Carol", "Alice", "Bob"]


def test_sort_descending(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name", "Score"], ["Carol", 30], ["Alice", 10], ["Bob", 20]]})

    response = _sort(file_id, {"sheet_name": "Data", "column": "B", "ascending": False}, commit=True)

    assert response.status_code == 200
    sorted_wb = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert [c.value for c in sorted_wb["Data"]["A"]] == ["Name", "Carol", "Bob", "Alice"]


def test_sort_without_header(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Carol", 30], ["Alice", 10], ["Bob", 20]]})

    response = _sort(
        file_id, {"sheet_name": "Data", "column": "B", "ascending": True, "has_header": False}, commit=True
    )

    assert response.status_code == 200
    body = response.json()
    assert body["row_count"] == 3
    sorted_wb = _load_stored_workbook(tmp_path, body["new_file_id"])
    # No pinned header -- row 1 (Carol/30) participates in the sort too.
    assert [c.value for c in sorted_wb["Data"]["A"]] == ["Alice", "Bob", "Carol"]


def test_sort_blanks_always_last_regardless_of_direction(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(
        {"Data": [["Name", "Score"], ["Alice", 10], ["Bob", None], ["Carol", 30]]}
    )

    ascending = _sort(file_id, {"sheet_name": "Data", "column": "B", "ascending": True}, commit=True)
    assert [c.value for c in _load_stored_workbook(tmp_path, ascending.json()["new_file_id"])["Data"]["A"]] == [
        "Name",
        "Alice",
        "Carol",
        "Bob",
    ]

    descending = _sort(file_id, {"sheet_name": "Data", "column": "B", "ascending": False}, commit=True)
    assert [c.value for c in _load_stored_workbook(tmp_path, descending.json()["new_file_id"])["Data"]["A"]] == [
        "Name",
        "Carol",
        "Alice",
        "Bob",
    ]


def test_sort_mixed_types_does_not_crash(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name", "Value"], ["A", 5], ["B", "banana"], ["C", 1], ["D", "apple"]]})

    response = _sort(file_id, {"sheet_name": "Data", "column": "B", "ascending": True}, commit=True)

    assert response.status_code == 200
    sorted_wb = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    # Numeric values sort together (1, 5), text values sort together (apple, banana),
    # numeric bucket ordered before text bucket.
    assert [c.value for c in sorted_wb["Data"]["A"]] == ["Name", "C", "A", "D", "B"]


def test_sort_untouched_sheets_pass_through(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets(
        {
            "Data": [["Name", "Score"], ["Bob", 20], ["Alice", 10]],
            "Other": [["X"], [3], [1], [2]],
        }
    )

    response = _sort(file_id, {"sheet_name": "Data", "column": "B", "ascending": True}, commit=True)

    assert response.status_code == 200
    sorted_wb = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert [c.value for c in sorted_wb["Other"]["A"]] == ["X", 3, 1, 2]


def test_sort_invalid_column_letters_returns_422(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _sort(file_id, {"sheet_name": "Data", "column": "2"})

    assert response.status_code == 422


def test_sort_invalid_sheet_name_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _sort(file_id, {"sheet_name": "DoesNotExist", "column": "A"})

    assert response.status_code == 400


def test_sort_column_out_of_bounds_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _sort(file_id, {"sheet_name": "Data", "column": "Z"})

    assert response.status_code == 400


def test_sort_missing_file_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    response = _sort("does-not-exist", {"sheet_name": "Data", "column": "A"})

    assert response.status_code == 404
