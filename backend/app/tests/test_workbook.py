import io

import openpyxl
from fastapi.testclient import TestClient

from app.core.config import get_settings
from app.main import app

client = TestClient(app)


def _make_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "Berlin"])
    ws.append(["Bob", 25, "Paris"])

    ws2 = wb.create_sheet("Empty")  # noqa: F841 - intentionally left blank

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _upload(content: bytes, filename: str = "test.xlsx") -> str:
    response = client.post(
        "/api/v1/upload",
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201
    return response.json()["file_id"]


def test_read_workbook_returns_structure(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload(_make_xlsx_bytes())

    response = client.get(f"/api/v1/workbook/{file_id}")

    assert response.status_code == 200
    body = response.json()
    assert body["file_id"] == file_id
    assert body["sheet_count"] == 2
    assert body["sheet_names"] == ["Sheet1", "Empty"]

    sheet1 = next(s for s in body["sheets"] if s["name"] == "Sheet1")
    assert sheet1["headers"] == ["Name", "Age", "City"]
    assert sheet1["preview_rows"] == [["Alice", 30, "Berlin"], ["Bob", 25, "Paris"]]
    assert sheet1["row_count"] == 3
    assert sheet1["column_count"] == 3

    empty_sheet = next(s for s in body["sheets"] if s["name"] == "Empty")
    assert empty_sheet["headers"] == []
    assert empty_sheet["preview_rows"] == []


def test_read_workbook_returns_structure_for_uppercase_extension(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    content = _make_xlsx_bytes()
    response = client.post(
        "/api/v1/upload",
        files={"file": ("test.XLSX", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    file_id = response.json()["file_id"]

    stored_files = list(tmp_path.glob(f"{file_id}*"))
    assert len(stored_files) == 1
    assert stored_files[0].suffix == ".xlsx"

    response = client.get(f"/api/v1/workbook/{file_id}")

    assert response.status_code == 200
    body = response.json()
    assert body["sheet_names"] == ["Sheet1", "Empty"]


def test_read_workbook_analysis_returns_column_stats(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload(_make_xlsx_bytes())

    response = client.get(f"/api/v1/workbook/{file_id}/analysis")

    assert response.status_code == 200
    body = response.json()
    assert body["file_id"] == file_id
    assert body["sheet_count"] == 2

    sheet1 = next(s for s in body["sheets"] if s["name"] == "Sheet1")
    columns = {c["name"]: c for c in sheet1["columns"]}

    assert columns["Name"]["inferred_type"] == "text"
    assert columns["Name"]["null_count"] == 0
    assert columns["Name"]["unique_count"] == 2
    assert columns["Name"]["min"] is None

    assert columns["Age"]["inferred_type"] == "numeric"
    assert columns["Age"]["null_count"] == 0
    assert columns["Age"]["unique_count"] == 2
    assert columns["Age"]["min"] == 25
    assert columns["Age"]["max"] == 30
    assert columns["Age"]["sum"] == 55
    assert columns["Age"]["mean"] == 27.5

    empty_sheet = next(s for s in body["sheets"] if s["name"] == "Empty")
    # openpyxl reports a brand-new blank sheet as having 1 column (its
    # default min_column/max_column), matching column_count in the
    # Workbook Reader's summary for the same sheet.
    assert [c["inferred_type"] for c in empty_sheet["columns"]] == ["empty"]


def test_read_workbook_analysis_handles_nulls_and_mixed_types(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Label", "Value"])
    ws.append(["a", 1])
    ws.append(["b", None])
    ws.append(["c", "not a number"])
    buffer = io.BytesIO()
    wb.save(buffer)

    file_id = _upload(buffer.getvalue())

    response = client.get(f"/api/v1/workbook/{file_id}/analysis")

    assert response.status_code == 200
    body = response.json()
    sheet1 = next(s for s in body["sheets"] if s["name"] == "Sheet1")
    columns = {c["name"]: c for c in sheet1["columns"]}

    value_col = columns["Value"]
    assert value_col["inferred_type"] == "mixed"
    assert value_col["null_count"] == 1
    assert value_col["unique_count"] == 2
    assert value_col["min"] is None
    assert value_col["mean"] is None


def test_read_workbook_returns_all_rows_uncapped(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["N"])
    for i in range(15):
        ws.append([i])
    buffer = io.BytesIO()
    wb.save(buffer)

    file_id = _upload(buffer.getvalue())

    response = client.get(f"/api/v1/workbook/{file_id}")

    assert response.status_code == 200
    sheet1 = next(s for s in response.json()["sheets"] if s["name"] == "Sheet1")
    assert len(sheet1["preview_rows"]) == 15
    assert sheet1["preview_rows"][-1] == [14]


def test_read_workbook_missing_file_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    response = client.get("/api/v1/workbook/does-not-exist")

    assert response.status_code == 404
