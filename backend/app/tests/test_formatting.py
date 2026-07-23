import io

import openpyxl
from fastapi.testclient import TestClient

from app.core.config import get_settings
from app.main import app

client = TestClient(app)


def _upload_sheets(sheets: dict[str, list[list]]) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)

    response = client.post(
        "/api/v1/upload",
        files={
            "file": (
                "test.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201
    return response.json()["file_id"]


def _format(file_id: str, request: dict, commit: bool = False):
    url = f"/api/v1/workbook/{file_id}/format"
    if commit:
        url += "?commit=true"
    return client.post(url, json=request)


def _load_stored_workbook(tmp_path, file_id: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(tmp_path / f"{file_id}.xlsx")


def test_format_header_row_preview_only(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name", "Age", "City"], ["Alice", 30, "Berlin"]]})

    response = _format(
        file_id,
        {"sheet_name": "Data", "header_row": True, "bold": True, "fill_color": "#4472C4"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["new_file_id"] is None
    assert body["sheet_name"] == "Data"
    assert body["range_applied"] == "A1:C1"
    assert body["cells_formatted"] == 3

    # Preview-only: the stored file must be untouched.
    original = _load_stored_workbook(tmp_path, file_id)
    header_cell = original["Data"]["A1"]
    assert header_cell.font.bold is not True
    assert header_cell.fill.fgColor.rgb in (None, "00000000")


def test_format_commit_writes_new_file_and_leaves_original_untouched(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name", "Age"], ["Alice", 30]]})

    response = _format(
        file_id,
        {"sheet_name": "Data", "header_row": True, "bold": True, "fill_color": "#4472C4"},
        commit=True,
    )

    assert response.status_code == 200
    body = response.json()
    new_file_id = body["new_file_id"]
    assert new_file_id is not None
    assert new_file_id != file_id

    cleaned = _load_stored_workbook(tmp_path, new_file_id)
    header_cell = cleaned["Data"]["A1"]
    assert header_cell.font.bold is True
    assert header_cell.fill.fgColor.rgb == "FF4472C4"
    # Formula-free data cell is untouched by the header-only range.
    assert cleaned["Data"]["A2"].font.bold is not True

    original = _load_stored_workbook(tmp_path, file_id)
    assert original["Data"]["A1"].font.bold is not True


def test_format_specific_range(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A", "B"], [1, 2], [3, 4]]})

    response = _format(file_id, {"sheet_name": "Data", "range": "B2:B3", "italic": True}, commit=True)

    assert response.status_code == 200
    body = response.json()
    assert body["range_applied"] == "B2:B3"
    assert body["cells_formatted"] == 2

    workbook = _load_stored_workbook(tmp_path, body["new_file_id"])
    assert workbook["Data"]["B2"].font.italic is True
    assert workbook["Data"]["B3"].font.italic is True
    assert workbook["Data"]["A2"].font.italic is not True


def test_format_number_format(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Rate"], [0.5]]})

    response = _format(
        file_id, {"sheet_name": "Data", "range": "A2:A2", "number_format": "0.00%"}, commit=True
    )

    assert response.status_code == 200
    workbook = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert workbook["Data"]["A2"].number_format == "0.00%"


def test_format_range_out_of_bounds_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _format(file_id, {"sheet_name": "Data", "range": "A1:Z50", "bold": True})

    assert response.status_code == 400


def test_format_invalid_sheet_name_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _format(file_id, {"sheet_name": "DoesNotExist", "bold": True})

    assert response.status_code == 400


def test_format_header_row_and_range_are_mutually_exclusive(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _format(file_id, {"sheet_name": "Data", "header_row": True, "range": "A1:A1", "bold": True})

    assert response.status_code == 422


def test_format_requires_at_least_one_style_property(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _format(file_id, {"sheet_name": "Data"})

    assert response.status_code == 422


def test_format_border_color_requires_border_style(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _format(file_id, {"sheet_name": "Data", "border_color": "#000000"})

    assert response.status_code == 422


def test_format_rejects_invalid_hex_color(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _format(file_id, {"sheet_name": "Data", "fill_color": "not-a-color"})

    assert response.status_code == 422


def test_format_border_style(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _format(
        file_id,
        {"sheet_name": "Data", "range": "A2:A2", "border_style": "thin", "border_color": "#FF0000"},
        commit=True,
    )

    assert response.status_code == 200
    workbook = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    cell = workbook["Data"]["A2"]
    assert cell.border.left.style == "thin"
    assert cell.border.left.color.rgb == "FFFF0000"


def test_format_missing_file_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    response = _format("does-not-exist", {"sheet_name": "Data", "bold": True})

    assert response.status_code == 404
