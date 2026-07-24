import io

import openpyxl
from fastapi.testclient import TestClient

from app.core.config import get_settings
from app.main import app

client = TestClient(app)


def _upload_sheets(sheets: dict[str, list[list]]) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)

    response = client.post(
        "/api/v1/upload",
        files={
            "file": (
                "test.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201
    return response.json()["file_id"]


def _rows_columns(file_id: str, request: dict, commit: bool = False):
    url = f"/api/v1/workbook/{file_id}/rows-columns"
    if commit:
        url += "?commit=true"
    return client.post(url, json=request)


def _load_stored_workbook(tmp_path, file_id: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(tmp_path / f"{file_id}.xlsx")


def test_insert_row_above_preview_only(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name", "Score"], ["Alice", 10], ["Bob", 20]]})

    response = _rows_columns(
        file_id, {"sheet_name": "Data", "action": "insert", "target": "row", "position": 2, "reference": "above"}
    )

    assert response.status_code == 200
    body = response.json()
    assert body["new_file_id"] is None
    assert body["new_row_count"] == 4

    original = _load_stored_workbook(tmp_path, file_id)
    assert original["Data"]["A2"].value == "Alice"


def test_insert_row_above_commit_shifts_rows_down(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name", "Score"], ["Alice", 10], ["Bob", 20]]})

    response = _rows_columns(
        file_id,
        {"sheet_name": "Data", "action": "insert", "target": "row", "position": 2, "reference": "above"},
        commit=True,
    )

    assert response.status_code == 200
    body = response.json()
    new_file_id = body["new_file_id"]
    assert new_file_id != file_id
    assert body["new_row_count"] == 4

    edited = _load_stored_workbook(tmp_path, new_file_id)
    assert edited["Data"]["A1"].value == "Name"
    assert edited["Data"]["A2"].value is None  # new blank row
    assert edited["Data"]["A3"].value == "Alice"
    assert edited["Data"]["A4"].value == "Bob"

    original = _load_stored_workbook(tmp_path, file_id)
    assert original["Data"]["A2"].value == "Alice"


def test_insert_row_below(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name"], ["Alice"], ["Bob"]]})

    response = _rows_columns(
        file_id,
        {"sheet_name": "Data", "action": "insert", "target": "row", "position": 2, "reference": "below"},
        commit=True,
    )

    assert response.status_code == 200
    edited = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert edited["Data"]["A1"].value == "Name"
    assert edited["Data"]["A2"].value == "Alice"
    assert edited["Data"]["A3"].value is None
    assert edited["Data"]["A4"].value == "Bob"


def test_insert_column_left_and_right(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A", "B"], [1, 2]]})

    left = _rows_columns(
        file_id,
        {"sheet_name": "Data", "action": "insert", "target": "column", "position": 2, "reference": "left"},
        commit=True,
    )
    assert left.status_code == 200
    left_wb = _load_stored_workbook(tmp_path, left.json()["new_file_id"])
    assert [c.value for c in left_wb["Data"][1]] == ["A", None, "B"]

    right = _rows_columns(
        file_id,
        {"sheet_name": "Data", "action": "insert", "target": "column", "position": 2, "reference": "right"},
        commit=True,
    )
    assert right.status_code == 200
    right_wb = _load_stored_workbook(tmp_path, right.json()["new_file_id"])
    assert [c.value for c in right_wb["Data"][1]] == ["A", "B", None]


def test_delete_row_single(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name"], ["Alice"], ["Bob"], ["Carol"]]})

    response = _rows_columns(
        file_id, {"sheet_name": "Data", "action": "delete", "target": "row", "position": 2}, commit=True
    )

    assert response.status_code == 200
    body = response.json()
    assert body["new_row_count"] == 3
    edited = _load_stored_workbook(tmp_path, body["new_file_id"])
    assert [c.value for c in edited["Data"]["A"]] == ["Name", "Bob", "Carol"]

    original = _load_stored_workbook(tmp_path, file_id)
    assert [c.value for c in original["Data"]["A"]] == ["Name", "Alice", "Bob", "Carol"]


def test_delete_row_multiple_via_count(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["Name"], ["Alice"], ["Bob"], ["Carol"], ["Dan"]]})

    response = _rows_columns(
        file_id,
        {"sheet_name": "Data", "action": "delete", "target": "row", "position": 2, "count": 2},
        commit=True,
    )

    assert response.status_code == 200
    edited = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert [c.value for c in edited["Data"]["A"]] == ["Name", "Carol", "Dan"]


def test_delete_column(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A", "B", "C"], [1, 2, 3]]})

    response = _rows_columns(
        file_id, {"sheet_name": "Data", "action": "delete", "target": "column", "position": 2}, commit=True
    )

    assert response.status_code == 200
    edited = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert [c.value for c in edited["Data"][1]] == ["A", "C"]


def test_invalid_sheet_name_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _rows_columns(
        file_id, {"sheet_name": "DoesNotExist", "action": "delete", "target": "row", "position": 1}
    )

    assert response.status_code == 400


def test_delete_position_out_of_bounds_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _rows_columns(file_id, {"sheet_name": "Data", "action": "delete", "target": "row", "position": 10})

    assert response.status_code == 400


def test_delete_count_extends_out_of_bounds_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1], [2]]})

    response = _rows_columns(
        file_id, {"sheet_name": "Data", "action": "delete", "target": "row", "position": 2, "count": 5}
    )

    assert response.status_code == 400


def test_insert_without_reference_returns_422(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _rows_columns(file_id, {"sheet_name": "Data", "action": "insert", "target": "row", "position": 1})

    assert response.status_code == 422


def test_insert_reference_mismatched_with_target_returns_422(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _rows_columns(
        file_id,
        {"sheet_name": "Data", "action": "insert", "target": "row", "position": 1, "reference": "left"},
    )

    assert response.status_code == 422


def test_delete_with_reference_returns_422(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _rows_columns(
        file_id,
        {"sheet_name": "Data", "action": "delete", "target": "row", "position": 1, "reference": "above"},
    )

    assert response.status_code == 422


def test_missing_file_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    response = _rows_columns(
        "does-not-exist", {"sheet_name": "Data", "action": "delete", "target": "row", "position": 1}
    )

    assert response.status_code == 404
