import io
import json

import openpyxl
from fastapi.testclient import TestClient

import app.services.ai_chat_service as ai_chat_service
from app.ai.provider import AIProvider
from app.core.config import get_settings
from app.main import app

client = TestClient(app)


class ScriptedAIProvider(AIProvider):
    """Returns each response in `responses` in order, one per call -- lets a
    test control exactly what the "AI" says without any real network call."""

    def __init__(self, *responses: str):
        self._responses = list(responses)

    async def generate_text(self, prompt: str) -> str:
        return self._responses.pop(0)


def _upload_sheets(sheets: dict[str, list[list]]) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)

    response = client.post(
        "/api/v1/upload",
        files={
            "file": (
                "test.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201
    return response.json()["file_id"]


def _chat(file_id: str, sheet_name: str, message: str, selection: str | None = None):
    return client.post(
        f"/api/v1/workbook/{file_id}/chat",
        json={"sheet_name": sheet_name, "selection": selection, "messages": [{"role": "user", "content": message}]},
    )


def _load_stored_workbook(tmp_path, file_id: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(tmp_path / f"{file_id}.xlsx")


def test_chat_reply_only_does_not_touch_file(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)
    ai_reply = json.dumps({"action": "reply", "message": "Bob had the highest score."})
    monkeypatch.setattr(ai_chat_service, "get_ai_provider", lambda: ScriptedAIProvider(ai_reply))

    file_id = _upload_sheets({"Data": [["Name", "Score"], ["Alice", 10], ["Bob", 90]]})

    response = _chat(file_id, "Data", "Who had the highest score?")

    assert response.status_code == 200
    body = response.json()
    assert body["reply"] == "Bob had the highest score."
    assert body["new_file_id"] is None
    assert body["engine"] is None


def test_chat_action_executes_and_commits(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)
    ai_reply = json.dumps(
        {
            "action": "format",
            "request": {"sheet_name": "Data", "header_row": True, "bold": True},
            "message": "Bolded the header row.",
        }
    )
    monkeypatch.setattr(ai_chat_service, "get_ai_provider", lambda: ScriptedAIProvider(ai_reply))

    file_id = _upload_sheets({"Data": [["Name", "Score"], ["Alice", 10]]})

    response = _chat(file_id, "Data", "Bold the header")

    assert response.status_code == 200
    body = response.json()
    assert body["reply"] == "Bolded the header row."
    assert body["engine"] == "format"
    new_file_id = body["new_file_id"]
    assert new_file_id is not None
    assert new_file_id != file_id

    edited = _load_stored_workbook(tmp_path, new_file_id)
    assert edited["Data"]["A1"].font.bold is True

    original = _load_stored_workbook(tmp_path, file_id)
    assert original["Data"]["A1"].font.bold is not True


def test_chat_action_can_be_rows_columns_engine(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)
    ai_reply = json.dumps(
        {
            "action": "rows_columns",
            "request": {"sheet_name": "Data", "action": "delete", "target": "row", "position": 2},
            "message": "Deleted row 2.",
        }
    )
    monkeypatch.setattr(ai_chat_service, "get_ai_provider", lambda: ScriptedAIProvider(ai_reply))

    file_id = _upload_sheets({"Data": [["Name"], ["Alice"], ["Bob"]]})

    response = _chat(file_id, "Data", "Delete row 2")

    assert response.status_code == 200
    body = response.json()
    assert body["engine"] == "rows_columns"
    edited = _load_stored_workbook(tmp_path, body["new_file_id"])
    assert [c.value for c in edited["Data"]["A"]] == ["Name", "Bob"]


def test_chat_falls_back_to_reply_on_invalid_json(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)
    monkeypatch.setattr(ai_chat_service, "get_ai_provider", lambda: ScriptedAIProvider("not json at all"))

    file_id = _upload_sheets({"Data": [["Name"], ["Alice"]]})

    response = _chat(file_id, "Data", "hello")

    assert response.status_code == 200
    body = response.json()
    assert body["reply"] == "not json at all"
    assert body["new_file_id"] is None


def test_chat_falls_back_to_reply_on_unknown_action(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)
    ai_reply = json.dumps({"action": "delete_everything", "message": "doing something scary"})
    monkeypatch.setattr(ai_chat_service, "get_ai_provider", lambda: ScriptedAIProvider(ai_reply))

    file_id = _upload_sheets({"Data": [["Name"], ["Alice"]]})

    response = _chat(file_id, "Data", "do something")

    assert response.status_code == 200
    body = response.json()
    assert body["new_file_id"] is None


def test_chat_falls_back_to_reply_on_schema_validation_failure(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)
    # "target" must be "row" or "column" -- this value fails schema validation.
    ai_reply = json.dumps(
        {
            "action": "rows_columns",
            "request": {"sheet_name": "Data", "action": "delete", "target": "planet", "position": 1},
            "message": "Deleting a planet.",
        }
    )
    monkeypatch.setattr(ai_chat_service, "get_ai_provider", lambda: ScriptedAIProvider(ai_reply))

    file_id = _upload_sheets({"Data": [["Name"], ["Alice"]]})

    response = _chat(file_id, "Data", "do something weird")

    assert response.status_code == 200
    body = response.json()
    assert body["new_file_id"] is None
    assert "rephrase" in body["reply"].lower()

    # Confirm nothing was written at all.
    original = _load_stored_workbook(tmp_path, file_id)
    assert [c.value for c in original["Data"]["A"]] == ["Name", "Alice"]


def test_chat_invalid_sheet_name_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)
    monkeypatch.setattr(ai_chat_service, "get_ai_provider", lambda: ScriptedAIProvider("unused"))

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _chat(file_id, "DoesNotExist", "hello")

    assert response.status_code == 400


def test_chat_missing_file_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)
    monkeypatch.setattr(ai_chat_service, "get_ai_provider", lambda: ScriptedAIProvider("unused"))

    response = _chat("does-not-exist", "Data", "hello")

    assert response.status_code == 404
