import io

import openpyxl
from fastapi.testclient import TestClient

from app.core.config import get_settings
from app.main import app

client = TestClient(app)


def _upload_sheets(sheets: dict[str, list[list]]) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)

    response = client.post(
        "/api/v1/upload",
        files={
            "file": (
                "test.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201
    return response.json()["file_id"]


def _apply(file_id: str, request: dict, commit: bool = False):
    url = f"/api/v1/workbook/{file_id}/formula"
    if commit:
        url += "?commit=true"
    return client.post(url, json=request)


def _load_stored_workbook(tmp_path, file_id: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(tmp_path / f"{file_id}.xlsx")


def test_formula_template_preview_only(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A", "B", "C"], [1, 2, 3], [4, 5, 6]]})

    response = _apply(file_id, {"sheet_name": "Data", "range": "D2:D3", "formula": "=SUM(A{row}:C{row})"})

    assert response.status_code == 200
    body = response.json()
    assert body["new_file_id"] is None
    assert body["range_applied"] == "D2:D3"
    assert body["cells_written"] == 2
    assert body["computed_value"] is None

    original = _load_stored_workbook(tmp_path, file_id)
    assert original["Data"]["D2"].value is None


def test_formula_template_commit_writes_relative_formulas(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A", "B", "C"], [1, 2, 3], [4, 5, 6]]})

    response = _apply(
        file_id, {"sheet_name": "Data", "range": "D2:D3", "formula": "=SUM(A{row}:C{row})"}, commit=True
    )

    assert response.status_code == 200
    new_file_id = response.json()["new_file_id"]
    assert new_file_id != file_id

    workbook = _load_stored_workbook(tmp_path, new_file_id)
    assert workbook["Data"]["D2"].value == "=SUM(A2:C2)"
    assert workbook["Data"]["D3"].value == "=SUM(A3:C3)"

    original = _load_stored_workbook(tmp_path, file_id)
    assert original["Data"]["D2"].value is None


def test_formula_template_col_placeholder(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A", "B", "C"], [1, 2, 3]]})

    response = _apply(file_id, {"sheet_name": "Data", "range": "A5:C5", "formula": "={col}1*2"}, commit=True)

    assert response.status_code == 200
    workbook = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert workbook["Data"]["A5"].value == "=A1*2"
    assert workbook["Data"]["B5"].value == "=B1*2"
    assert workbook["Data"]["C5"].value == "=C1*2"


def test_formula_template_can_write_beyond_current_bounds(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1], [2], [3]]})

    response = _apply(file_id, {"sheet_name": "Data", "range": "D1:D5", "formula": "=A{row}"}, commit=True)

    assert response.status_code == 200
    body = response.json()
    assert body["cells_written"] == 5

    workbook = _load_stored_workbook(tmp_path, body["new_file_id"])
    assert workbook["Data"]["D5"].value == "=A5"


def test_formula_template_invalid_range_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _apply(file_id, {"sheet_name": "Data", "range": "not-a-range", "formula": "=A1"})

    assert response.status_code == 400


def test_formula_function_sum_preview_only(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [[1], [2], [3]]})

    response = _apply(file_id, {"sheet_name": "Data", "cell": "B1", "function": "SUM", "source_range": "A1:A3"})

    assert response.status_code == 200
    body = response.json()
    assert body["new_file_id"] is None
    assert body["range_applied"] == "B1"
    assert body["cells_written"] == 1
    assert body["computed_value"] == 6

    original = _load_stored_workbook(tmp_path, file_id)
    assert original["Data"]["B1"].value is None


def test_formula_function_commit_writes_formula_and_leaves_original_untouched(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [[1], [2], [3]]})

    response = _apply(
        file_id, {"sheet_name": "Data", "cell": "B1", "function": "SUM", "source_range": "A1:A3"}, commit=True
    )

    assert response.status_code == 200
    workbook = _load_stored_workbook(tmp_path, response.json()["new_file_id"])
    assert workbook["Data"]["B1"].value == "=SUM(A1:A3)"

    original = _load_stored_workbook(tmp_path, file_id)
    assert original["Data"]["B1"].value is None


def test_formula_function_average(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [[2], [4], [6]]})

    response = _apply(file_id, {"sheet_name": "Data", "cell": "B1", "function": "AVERAGE", "source_range": "A1:A3"})

    assert response.status_code == 200
    assert response.json()["computed_value"] == 4


def test_formula_function_count_ignores_non_numeric(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [[1], ["text"], [3]]})

    response = _apply(file_id, {"sheet_name": "Data", "cell": "B1", "function": "COUNT", "source_range": "A1:A3"})

    assert response.status_code == 200
    assert response.json()["computed_value"] == 2


def test_formula_function_min_max(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [[5], [1], [9]]})

    min_response = _apply(file_id, {"sheet_name": "Data", "cell": "B1", "function": "MIN", "source_range": "A1:A3"})
    max_response = _apply(file_id, {"sheet_name": "Data", "cell": "C1", "function": "MAX", "source_range": "A1:A3"})

    assert min_response.json()["computed_value"] == 1
    assert max_response.json()["computed_value"] == 9


def test_formula_function_average_with_no_numeric_values_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], ["x"], ["y"]]})

    response = _apply(file_id, {"sheet_name": "Data", "cell": "B1", "function": "AVERAGE", "source_range": "A1:A2"})

    assert response.status_code == 400


def test_formula_function_rejects_uncomputed_formula_in_source(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["A"])
    ws.append([1])
    ws.append(["=A2+1"])
    buffer = io.BytesIO()
    wb.save(buffer)

    response = client.post(
        "/api/v1/upload",
        files={
            "file": (
                "test.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    file_id = response.json()["file_id"]

    result = _apply(file_id, {"sheet_name": "Data", "cell": "B1", "function": "SUM", "source_range": "A1:A3"})

    assert result.status_code == 400


def test_formula_function_source_range_out_of_bounds_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _apply(file_id, {"sheet_name": "Data", "cell": "B1", "function": "SUM", "source_range": "A1:A50"})

    assert response.status_code == 400


def test_formula_invalid_sheet_name_returns_400(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _apply(file_id, {"sheet_name": "DoesNotExist", "range": "B1:B1", "formula": "=A1"})

    assert response.status_code == 400


def test_formula_rejects_mixing_both_modes(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _apply(
        file_id,
        {
            "sheet_name": "Data",
            "range": "B1:B1",
            "formula": "=A1",
            "cell": "C1",
            "function": "SUM",
            "source_range": "A1:A1",
        },
    )

    assert response.status_code == 422


def test_formula_requires_a_mode(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _apply(file_id, {"sheet_name": "Data"})

    assert response.status_code == 422


def test_formula_must_start_with_equals(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    file_id = _upload_sheets({"Data": [["A"], [1]]})

    response = _apply(file_id, {"sheet_name": "Data", "range": "B1:B1", "formula": "SUM(A1)"})

    assert response.status_code == 422


def test_formula_missing_file_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(get_settings(), "STORAGE_DIR", tmp_path)

    response = _apply("does-not-exist", {"sheet_name": "Data", "range": "B1:B1", "formula": "=A1"})

    assert response.status_code == 404
