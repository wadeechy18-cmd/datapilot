"""
Insights engine.

Computes real, deterministic statistical findings for one sheet -- outliers
(IQR method), exact-duplicate rows, per-column trends (correlation against
row order), and correlations between numeric column pairs. Entirely local:
plain Python + the stdlib `statistics` module, no numpy/pandas, no AI.
Read-only, like the reader/analyzer -- it opens the workbook independently
and never writes anything.

Only genuinely notable findings are reported (a column with no outliers,
a weak trend, or a weak correlation is simply omitted) so the output stays
a short list of real signals, not noise -- both for direct display in the
UI and for feeding into the AI summary/chat prompts as real computed
context instead of just column aggregates.
"""

import statistics
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.insights import ColumnCorrelation, ColumnOutliers, ColumnTrend, SheetInsights

_TREND_THRESHOLD = 0.7
_CORRELATION_THRESHOLD = 0.7
_MAX_SAMPLE_OUTLIERS = 5


class SheetNotFoundError(Exception):
    """Raised when the requested sheet_name doesn't exist in the workbook."""


def _read_sheet_raw(worksheet: Worksheet) -> tuple[list[Any], list[list[Any]]]:
    rows_iter = worksheet.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    headers = list(first_row) if first_row is not None else []
    data_rows = [list(row) for row in rows_iter]
    return headers, data_rows


def _numeric_column(data_rows: list[list[Any]], col_index: int) -> list[float]:
    values = []
    for row in data_rows:
        value = row[col_index] if col_index < len(row) else None
        if isinstance(value, bool) or value is None:
            continue
        if isinstance(value, (int, float)):
            values.append(float(value))
    return values


def _pearson_correlation(xs: list[float], ys: list[float]) -> float | None:
    n = len(xs)
    if n < 3:
        return None
    mean_x = sum(xs) / n
    mean_y = sum(ys) / n
    covariance = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    variance_x = sum((x - mean_x) ** 2 for x in xs)
    variance_y = sum((y - mean_y) ** 2 for y in ys)
    if variance_x == 0 or variance_y == 0:
        return None
    return covariance / (variance_x**0.5 * variance_y**0.5)


def _detect_outliers(column_name: Any, values: list[float]) -> ColumnOutliers | None:
    if len(values) < 4:
        return None
    q1, _, q3 = statistics.quantiles(values, n=4, method="inclusive")
    iqr = q3 - q1
    if iqr == 0:
        return None
    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr
    outliers = sorted((v for v in values if v < lower_bound or v > upper_bound))
    if not outliers:
        return None
    return ColumnOutliers(
        column=column_name,
        outlier_count=len(outliers),
        lower_bound=round(lower_bound, 2),
        upper_bound=round(upper_bound, 2),
        sample_values=outliers[:_MAX_SAMPLE_OUTLIERS],
    )


def _detect_trend(column_name: Any, values: list[float]) -> ColumnTrend | None:
    if len(values) < 3:
        return None
    row_order = [float(i) for i in range(len(values))]
    r = _pearson_correlation(row_order, values)
    if r is None or abs(r) < _TREND_THRESHOLD:
        return None
    direction = "increasing" if r > 0 else "decreasing"
    return ColumnTrend(column=column_name, direction=direction, strength=round(r, 3))


def _count_duplicate_rows(data_rows: list[list[Any]]) -> int:
    seen: set[tuple[Any, ...]] = set()
    duplicate_count = 0
    for row in data_rows:
        key = tuple(row)
        if key in seen:
            duplicate_count += 1
        else:
            seen.add(key)
    return duplicate_count


def compute_insights(file_id: str, path: Path, sheet_name: str | None) -> SheetInsights:
    """Load the workbook at `path` and compute insights for one sheet
    (`sheet_name`, or the workbook's first sheet if omitted). Does not
    write anything to disk."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        target_name = sheet_name or (sheet_names[0] if sheet_names else None)
        if target_name is None or target_name not in sheet_names:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in workbook '{file_id}'.")

        headers, data_rows = _read_sheet_raw(workbook[target_name])
    finally:
        workbook.close()

    numeric_columns: dict[int, list[float]] = {}
    for index, header in enumerate(headers):
        values = _numeric_column(data_rows, index)
        if len(values) >= 3:
            numeric_columns[index] = values

    outliers = []
    trends = []
    for index, values in numeric_columns.items():
        column_name = headers[index]
        outlier = _detect_outliers(column_name, values)
        if outlier is not None:
            outliers.append(outlier)
        trend = _detect_trend(column_name, values)
        if trend is not None:
            trends.append(trend)

    correlations = []
    column_indices = sorted(numeric_columns.keys())
    for i, index_a in enumerate(column_indices):
        for index_b in column_indices[i + 1 :]:
            values_a, values_b = numeric_columns[index_a], numeric_columns[index_b]
            paired_length = min(len(values_a), len(values_b))
            r = _pearson_correlation(values_a[:paired_length], values_b[:paired_length])
            if r is not None and abs(r) >= _CORRELATION_THRESHOLD:
                correlations.append(
                    ColumnCorrelation(column_a=headers[index_a], column_b=headers[index_b], correlation=round(r, 3))
                )

    return SheetInsights(
        file_id=file_id,
        sheet_name=target_name,
        duplicate_row_count=_count_duplicate_rows(data_rows),
        outliers=outliers,
        trends=trends,
        correlations=correlations,
    )


def format_insights_for_prompt(insights: SheetInsights) -> str:
    """Renders computed insights as compact text for an AI prompt -- real,
    computed findings, not the AI's own guesswork."""
    lines: list[str] = []
    if insights.duplicate_row_count > 0:
        lines.append(f"- {insights.duplicate_row_count} exact-duplicate row(s).")
    for outlier in insights.outliers:
        lines.append(
            f"- Column {outlier.column!r} has {outlier.outlier_count} statistical outlier(s) "
            f"(expected range {outlier.lower_bound}-{outlier.upper_bound}), e.g. {outlier.sample_values}."
        )
    for trend in insights.trends:
        lines.append(f"- Column {trend.column!r} trends {trend.direction} down the sheet (r={trend.strength}).")
    for correlation in insights.correlations:
        lines.append(
            f"- Columns {correlation.column_a!r} and {correlation.column_b!r} are strongly correlated "
            f"(r={correlation.correlation})."
        )
    if not lines:
        return "No notable outliers, duplicate rows, trends, or correlations were found."
    return "\n".join(lines)
