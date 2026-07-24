"""
Rows & columns structural-edit engine.

Inserts or deletes whole rows/columns on a sheet, using openpyxl's native
insert_rows/insert_cols/delete_rows/delete_cols. Like the formatting engine,
this only mutates an in-memory copy of the workbook -- nothing is written to
disk unless the caller commits, and committing always writes a new file
rather than overwriting the original. Loads *without* data_only, to keep
existing formula strings intact.

Known limitation: openpyxl does not rewrite formula references when rows or
columns shift (e.g. a "=A1" formula below an inserted row still reads "=A1",
not "=A2"). This is an openpyxl limitation, not something this engine works
around.
"""

import io
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook

from app.models.rows_columns import RowColumnResult
from app.schemas.rows_columns import RowColumnRequest


class SheetNotFoundError(Exception):
    """Raised when the requested sheet_name doesn't exist in the workbook."""


class InvalidPositionError(Exception):
    """Raised when position/count fall outside the sheet's current bounds."""


def _resolve_insert_index(request: RowColumnRequest) -> int:
    """The 1-based index openpyxl should insert *before*."""
    if request.reference in ("below", "right"):
        return request.position + 1
    return request.position


def apply_row_column_operation(
    file_id: str, path: Path, request: RowColumnRequest
) -> tuple[RowColumnResult, Workbook]:
    """Load the workbook at `path`, insert/delete the requested rows or
    columns in memory, and return both a summary and the mutated Workbook.
    The caller decides whether to persist it (see write_rows_columns_workbook)
    -- nothing here touches the file on disk."""
    workbook = openpyxl.load_workbook(path)
    if request.sheet_name not in workbook.sheetnames:
        raise SheetNotFoundError(f"Sheet '{request.sheet_name}' not found in workbook '{file_id}'.")

    worksheet = workbook[request.sheet_name]
    max_row = worksheet.max_row or 1
    max_col = worksheet.max_column or 1
    bound = max_row if request.target == "row" else max_col
    bound_label = "row(s)" if request.target == "row" else "column(s)"

    if request.action == "delete":
        if request.position > bound or request.position + request.count - 1 > bound:
            raise InvalidPositionError(
                f"Sheet '{request.sheet_name}' only has {bound} {bound_label}; "
                f"can't delete position {request.position} with count {request.count}."
            )
        if request.target == "row":
            worksheet.delete_rows(request.position, request.count)
        else:
            worksheet.delete_cols(request.position, request.count)
    else:
        if request.position > bound:
            raise InvalidPositionError(
                f"Sheet '{request.sheet_name}' only has {bound} {bound_label}; "
                f"can't insert relative to position {request.position}."
            )
        insert_index = _resolve_insert_index(request)
        if request.target == "row":
            worksheet.insert_rows(insert_index, request.count)
            # insert_rows() only shifts cells that already exist below the
            # insertion point -- if there's nothing there (e.g. inserting
            # "below" the sheet's last row), it's a silent no-op and the new
            # row never materializes, since openpyxl drops genuinely empty
            # (value=None) cells on save. Writing "" instead of leaving the
            # cell untouched forces it to persist -- and still reads back as
            # a blank (None) value everywhere else in the app.
            for offset in range(request.count):
                for col in range(1, max_col + 1):
                    worksheet.cell(row=insert_index + offset, column=col, value="")
        else:
            worksheet.insert_cols(insert_index, request.count)
            for offset in range(request.count):
                for row in range(1, max_row + 1):
                    worksheet.cell(row=row, column=insert_index + offset, value="")

    result = RowColumnResult(
        file_id=file_id,
        sheet_name=request.sheet_name,
        action=request.action,
        target=request.target,
        position=request.position,
        count=request.count,
        new_row_count=worksheet.max_row or 0,
        new_column_count=worksheet.max_column or 0,
    )
    return result, workbook


def write_rows_columns_workbook(workbook: Workbook) -> bytes:
    """Serialize a (mutated, in-memory) workbook to .xlsx bytes, for committing to storage."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
