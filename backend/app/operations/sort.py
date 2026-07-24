"""
Sort engine.

Reorders a sheet's data rows by one column's value, optionally pinning a
header row (has_header) so it never moves. openpyxl has no native row-sort,
so -- like the cleaning engine -- this rebuilds a fresh workbook from read
values rather than mutating in place. That means, exactly like a Cleaning
commit today, cell styles and formulas on data rows are NOT preserved
through a sort commit; only values are. Read-only with respect to the
stored file -- it opens the workbook independently and returns the sorted
data in memory; writing the result to disk (write_sorted_workbook) is a
separate step the caller opts into.
"""

import io
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from app.models.sort import SheetData, WorkbookSortResult
from app.schemas.sort import SortRequest


class SheetNotFoundError(Exception):
    """Raised when the requested sheet_name doesn't exist in the workbook."""


class InvalidColumnError(Exception):
    """Raised when the requested sort column is out of the sheet's bounds."""


def _read_sheet_raw(worksheet: Worksheet) -> tuple[list[Any], list[list[Any]]]:
    rows_iter = worksheet.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    headers = list(first_row) if first_row is not None else []
    data_rows = [list(row) for row in rows_iter]
    return headers, data_rows


def _sort_key(value: Any) -> tuple[int, Any]:
    """Groups values by comparable type so a mixed-type column never raises
    a TypeError when compared (Python tuple comparison only looks at the
    second element once the first elements are equal, so numeric and text
    ranks are never compared against each other directly)."""
    if isinstance(value, bool):
        return (1, str(value).lower())
    if isinstance(value, (int, float)):
        return (0, value)
    if isinstance(value, str):
        try:
            return (0, float(value))
        except ValueError:
            return (1, value.lower())
    return (1, str(value))


def sort_workbook(file_id: str, path: Path, request: SortRequest) -> WorkbookSortResult:
    """Load the workbook at `path` and return every sheet's data, with
    `request.sheet_name`'s data rows reordered by `request.column`. Blank
    cells in the sort column always sort last, regardless of direction.
    Does not write anything to disk."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        if request.sheet_name not in sheet_names:
            raise SheetNotFoundError(f"Sheet '{request.sheet_name}' not found in workbook '{file_id}'.")

        column_index = column_index_from_string(request.column) - 1

        sheets: list[SheetData] = []
        sorted_row_count = 0
        for name in sheet_names:
            worksheet = workbook[name]

            if name != request.sheet_name:
                headers, data_rows = _read_sheet_raw(worksheet)
                sheets.append(SheetData(name=name, headers=headers, rows=data_rows))
                continue

            all_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if request.has_header and all_rows:
                headers, data_rows = all_rows[0], all_rows[1:]
            else:
                headers, data_rows = [], all_rows

            max_columns = max((len(row) for row in all_rows), default=0)
            if column_index < 0 or column_index >= max_columns:
                raise InvalidColumnError(
                    f"Column '{request.column}' is out of bounds for sheet '{name}' ({max_columns} column(s))."
                )

            def key_fn(row: list[Any], column_index: int = column_index) -> Any:
                value = row[column_index] if column_index < len(row) else None
                return _sort_key(value)

            def is_blank(row: list[Any], column_index: int = column_index) -> bool:
                value = row[column_index] if column_index < len(row) else None
                return value is None or value == ""

            blanks = [row for row in data_rows if is_blank(row)]
            non_blanks = [row for row in data_rows if not is_blank(row)]
            non_blanks.sort(key=key_fn, reverse=not request.ascending)
            data_rows = non_blanks + blanks
            sorted_row_count = len(data_rows)

            sheets.append(SheetData(name=name, headers=headers, rows=data_rows))
    finally:
        workbook.close()

    return WorkbookSortResult(
        file_id=file_id,
        sheet_name=request.sheet_name,
        column=request.column,
        ascending=request.ascending,
        has_header=request.has_header,
        row_count=sorted_row_count,
        sheets=sheets,
    )


def write_sorted_workbook(result: WorkbookSortResult) -> bytes:
    """Serialize a sort result to .xlsx bytes, for committing to storage."""
    output = openpyxl.Workbook()
    output.remove(output.active)

    for sheet in result.sheets:
        worksheet = output.create_sheet(title=sheet.name)
        if sheet.headers:
            worksheet.append(sheet.headers)
        for row in sheet.rows:
            worksheet.append(row)

    buffer = io.BytesIO()
    output.save(buffer)
    return buffer.getvalue()
