"""
Cleaning engine.

Computes a cleaned version of a workbook's sheet(s): trims whitespace,
drops fully-empty rows/columns, removes exact duplicate rows, and
optionally fills or drops nulls. Read-only with respect to the stored
file — it opens the workbook independently and returns the cleaned data
in memory; writing the result to disk (write_cleaned_workbook) is a
separate step the caller opts into.
"""

import io
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.cleaning import CleanedSheet, WorkbookCleaningResult
from app.schemas.cleaning import CleaningRequest

_NO_FILL = object()


class SheetNotFoundError(Exception):
    """Raised when a requested sheet_name doesn't exist in the workbook."""


def _compute_fill_value(strategy: str, non_null_values: list[Any], placeholder: Any) -> Any:
    """Returns the value to fill nulls with, or _NO_FILL if this strategy
    doesn't apply to this column (e.g. "mean" on a non-numeric column)."""
    if strategy == "placeholder":
        return placeholder
    if strategy == "mode":
        if not non_null_values:
            return _NO_FILL
        return Counter(non_null_values).most_common(1)[0][0]
    if strategy in ("zero", "mean"):
        numeric_values = [v for v in non_null_values if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if not numeric_values or len(numeric_values) != len(non_null_values):
            return _NO_FILL
        return 0 if strategy == "zero" else sum(numeric_values) / len(numeric_values)
    return _NO_FILL


def _clean_rows(
    headers: list[Any], rows: list[list[Any]], request: CleaningRequest
) -> tuple[list[Any], list[list[Any]], dict[str, int]]:
    headers = list(headers)
    working_rows = [list(row) for row in rows]
    cells_trimmed = 0

    if request.trim_whitespace:
        headers = [h.strip() if isinstance(h, str) else h for h in headers]
        for row in working_rows:
            for i, value in enumerate(row):
                if isinstance(value, str):
                    trimmed = value.strip()
                    if trimmed != value:
                        cells_trimmed += 1
                    row[i] = trimmed or None

    columns_removed = 0
    if request.drop_empty_columns and headers:
        keep_indices = [i for i in range(len(headers)) if any(row[i] is not None for row in working_rows)]
        columns_removed = len(headers) - len(keep_indices)
        headers = [headers[i] for i in keep_indices]
        working_rows = [[row[i] for i in keep_indices] for row in working_rows]

    if request.drop_empty_rows:
        working_rows = [row for row in working_rows if any(v is not None for v in row)]

    if request.drop_duplicate_rows:
        seen: set[tuple[Any, ...]] = set()
        deduped: list[list[Any]] = []
        for row in working_rows:
            key = tuple(row)
            if key not in seen:
                seen.add(key)
                deduped.append(row)
        working_rows = deduped

    nulls_filled = 0
    if request.drop_rows_with_nulls:
        working_rows = [row for row in working_rows if all(v is not None for v in row)]
    elif request.fill_nulls is not None:
        strategy = request.fill_nulls.strategy
        placeholder = request.fill_nulls.placeholder
        for col_index in range(len(headers)):
            column_values = [row[col_index] for row in working_rows]
            non_null_values = [v for v in column_values if v is not None]
            if len(non_null_values) == len(column_values):
                continue
            fill_value = _compute_fill_value(strategy, non_null_values, placeholder)
            if fill_value is _NO_FILL:
                continue
            for row in working_rows:
                if row[col_index] is None:
                    row[col_index] = fill_value
                    nulls_filled += 1

    stats = {
        "columns_removed": columns_removed,
        "cells_trimmed": cells_trimmed,
        "nulls_filled": nulls_filled,
    }
    return headers, working_rows, stats


def _read_sheet_raw(worksheet: Worksheet) -> tuple[list[Any], list[list[Any]]]:
    rows_iter = worksheet.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    headers = list(first_row) if first_row is not None else []
    data_rows = [list(row) for row in rows_iter]
    return headers, data_rows


def clean_workbook(file_id: str, path: Path, request: CleaningRequest) -> WorkbookCleaningResult:
    """Load the workbook at `path` and return the cleaned data for each
    sheet. Sheets other than `request.sheet_name` (if set) pass through
    unchanged. Does not write anything to disk."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        if request.sheet_name is not None and request.sheet_name not in sheet_names:
            raise SheetNotFoundError(f"Sheet '{request.sheet_name}' not found in workbook '{file_id}'.")

        sheets: list[CleanedSheet] = []
        for name in sheet_names:
            headers, data_rows = _read_sheet_raw(workbook[name])
            original_row_count = len(data_rows)
            original_column_count = len(headers)

            if request.sheet_name is None or request.sheet_name == name:
                cleaned_headers, cleaned_rows, stats = _clean_rows(headers, data_rows, request)
            else:
                cleaned_headers, cleaned_rows = headers, data_rows
                stats = {"columns_removed": 0, "cells_trimmed": 0, "nulls_filled": 0}

            sheets.append(
                CleanedSheet(
                    name=name,
                    original_row_count=original_row_count,
                    cleaned_row_count=len(cleaned_rows),
                    original_column_count=original_column_count,
                    cleaned_column_count=len(cleaned_headers),
                    rows_removed=original_row_count - len(cleaned_rows),
                    columns_removed=stats["columns_removed"],
                    cells_trimmed=stats["cells_trimmed"],
                    nulls_filled=stats["nulls_filled"],
                    headers=cleaned_headers,
                    rows=cleaned_rows,
                )
            )
    finally:
        workbook.close()

    return WorkbookCleaningResult(file_id=file_id, sheets=sheets)


def write_cleaned_workbook(result: WorkbookCleaningResult) -> bytes:
    """Serialize a cleaning result to .xlsx bytes, for committing to storage."""
    output = openpyxl.Workbook()
    output.remove(output.active)

    for sheet in result.sheets:
        worksheet = output.create_sheet(title=sheet.name)
        if sheet.headers:
            worksheet.append(sheet.headers)
        for row in sheet.rows:
            worksheet.append(row)

    buffer = io.BytesIO()
    output.save(buffer)
    return buffer.getvalue()
