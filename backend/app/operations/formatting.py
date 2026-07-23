"""
Formatting engine.

Applies cell-level styling (font, fill, number format, alignment,
borders) to a sheet's whole used range, a specific cell range, or just
the header row. Like the cleaning engine, this only mutates an in-memory
copy of the workbook -- nothing is written to disk unless the caller
commits, and committing always writes a new file rather than overwriting
the original. Unlike the reader/analyzer/cleaner, this loads the workbook
*without* data_only, since formulas must be preserved exactly -- only
cell styles are touched, never values.
"""

import io
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.formatting import FormattingResult
from app.schemas.formatting import FormattingRequest


class SheetNotFoundError(Exception):
    """Raised when the requested sheet_name doesn't exist in the workbook."""


class InvalidRangeError(Exception):
    """Raised when the requested range string can't be parsed or is out of bounds."""


def _normalize_color(value: str) -> str:
    """openpyxl expects an 8-digit ARGB hex string; accept '#RRGGBB' or 'RRGGBB' too."""
    hex_value = value.lstrip("#").upper()
    return hex_value if len(hex_value) == 8 else "FF" + hex_value


def _resolve_bounds(worksheet: Worksheet, request: FormattingRequest) -> tuple[int, int, int, int]:
    max_row = worksheet.max_row or 1
    max_col = worksheet.max_column or 1

    if request.header_row:
        return (1, 1, max_col, 1)

    if request.range is not None:
        try:
            min_col, min_row, range_max_col, range_max_row = range_boundaries(request.range)
        except ValueError as exc:
            raise InvalidRangeError(f"Invalid range '{request.range}': {exc}") from exc
        if min_row < 1 or min_col < 1 or range_max_row > max_row or range_max_col > max_col:
            raise InvalidRangeError(
                f"Range '{request.range}' is out of bounds for a sheet with "
                f"{max_row} row(s) and {max_col} column(s)."
            )
        return (min_col, min_row, range_max_col, range_max_row)

    return (1, 1, max_col, max_row)


def _apply_style(cell, request: FormattingRequest) -> None:
    if (
        request.bold is not None
        or request.italic is not None
        or request.font_size is not None
        or request.font_color is not None
    ):
        existing = cell.font
        cell.font = Font(
            name=existing.name,
            bold=request.bold if request.bold is not None else existing.bold,
            italic=request.italic if request.italic is not None else existing.italic,
            size=request.font_size if request.font_size is not None else existing.size,
            color=_normalize_color(request.font_color) if request.font_color is not None else existing.color,
        )

    if request.fill_color is not None:
        color = _normalize_color(request.fill_color)
        cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)

    if request.number_format is not None:
        cell.number_format = request.number_format

    if request.horizontal_alignment is not None or request.vertical_alignment is not None:
        existing_alignment = cell.alignment
        cell.alignment = Alignment(
            horizontal=(
                request.horizontal_alignment
                if request.horizontal_alignment is not None
                else existing_alignment.horizontal
            ),
            vertical=(
                request.vertical_alignment if request.vertical_alignment is not None else existing_alignment.vertical
            ),
        )

    if request.border_style is not None:
        color = _normalize_color(request.border_color) if request.border_color is not None else "FF000000"
        side = Side(style=request.border_style, color=color)
        cell.border = Border(left=side, right=side, top=side, bottom=side)


def apply_formatting(file_id: str, path: Path, request: FormattingRequest) -> tuple[FormattingResult, Workbook]:
    """Load the workbook at `path`, apply the requested styling in memory,
    and return both a summary of what changed and the mutated Workbook.
    The caller decides whether to persist it (see write_formatted_workbook) --
    nothing here touches the file on disk."""
    workbook = openpyxl.load_workbook(path)
    if request.sheet_name not in workbook.sheetnames:
        raise SheetNotFoundError(f"Sheet '{request.sheet_name}' not found in workbook '{file_id}'.")

    worksheet = workbook[request.sheet_name]
    min_col, min_row, max_col, max_row = _resolve_bounds(worksheet, request)

    cells_formatted = 0
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            _apply_style(cell, request)
            cells_formatted += 1

    range_applied = f"{worksheet.cell(min_row, min_col).coordinate}:{worksheet.cell(max_row, max_col).coordinate}"

    result = FormattingResult(
        file_id=file_id,
        sheet_name=request.sheet_name,
        range_applied=range_applied,
        cells_formatted=cells_formatted,
    )
    return result, workbook


def write_formatted_workbook(workbook: Workbook) -> bytes:
    """Serialize a (mutated, in-memory) workbook to .xlsx bytes, for committing to storage."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
