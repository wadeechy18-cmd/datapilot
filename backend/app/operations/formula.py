"""
Formula engine.

Writes formulas into a workbook's cells, in one of two modes:

- Template: a formula string applied across a destination range, with
  {row}/{col} substituted per cell (like Excel's "fill down"). The
  destination is allowed to extend beyond the sheet's current used
  range -- adding a new formula column/row is the main use case -- so
  it's only checked against Excel's absolute sheet limits, not the
  sheet's current dimensions.
- Function: a convenience aggregate (SUM/AVERAGE/COUNT/MIN/MAX) written
  to a single destination cell, whose result is also computed locally
  over literal values in a source range. The source range IS checked
  against the sheet's current dimensions, since it must reference real
  data; a source cell holding an uncomputed formula can't be evaluated
  locally and is rejected with a clear error.

Like the cleaning/formatting engines, this only mutates an in-memory
copy of the workbook -- nothing is written to disk unless the caller
commits, and committing always writes a new file rather than
overwriting the original. Loaded *without* data_only, since existing
formulas in the workbook must be preserved exactly.
"""

import io
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string, range_boundaries
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.formula import FormulaResult
from app.schemas.formula import FormulaRequest

_EXCEL_MAX_ROW = 1_048_576
_EXCEL_MAX_COL = 16_384


class SheetNotFoundError(Exception):
    """Raised when the requested sheet_name doesn't exist in the workbook."""


class InvalidRangeError(Exception):
    """Raised when a range/cell reference can't be parsed, is out of bounds,
    or (for aggregate functions) contains an uncomputed formula that can't
    be evaluated locally."""


def _validate_write_bounds(min_col: int, min_row: int, max_col: int, max_row: int, ref: str) -> None:
    if min_row < 1 or min_col < 1:
        raise InvalidRangeError(f"Invalid range '{ref}': row/column must be 1 or greater.")
    if max_row > _EXCEL_MAX_ROW or max_col > _EXCEL_MAX_COL:
        raise InvalidRangeError(f"Range '{ref}' exceeds Excel's maximum sheet size.")


def _validate_read_bounds(worksheet: Worksheet, min_col: int, min_row: int, max_col: int, max_row: int, ref: str) -> None:
    sheet_max_row = worksheet.max_row or 1
    sheet_max_col = worksheet.max_column or 1
    if min_row < 1 or min_col < 1 or max_row > sheet_max_row or max_col > sheet_max_col:
        raise InvalidRangeError(
            f"Range '{ref}' is out of bounds for a sheet with {sheet_max_row} row(s) and {sheet_max_col} column(s)."
        )


def _parse_range_for_write(ref: str) -> tuple[int, int, int, int]:
    try:
        bounds = range_boundaries(ref)
    except ValueError as exc:
        raise InvalidRangeError(f"Invalid range '{ref}': {exc}") from exc
    _validate_write_bounds(*bounds, ref)
    return bounds


def _parse_range_for_read(worksheet: Worksheet, ref: str) -> tuple[int, int, int, int]:
    try:
        bounds = range_boundaries(ref)
    except ValueError as exc:
        raise InvalidRangeError(f"Invalid range '{ref}': {exc}") from exc
    _validate_read_bounds(worksheet, *bounds, ref)
    return bounds


def _parse_cell_for_write(ref: str) -> tuple[int, int]:
    try:
        column_letter, row = coordinate_from_string(ref)
    except ValueError as exc:
        raise InvalidRangeError(f"Invalid cell reference '{ref}': {exc}") from exc
    col = column_index_from_string(column_letter)
    _validate_write_bounds(col, row, col, row, ref)
    return col, row


def _apply_template(worksheet: Worksheet, request: FormulaRequest) -> tuple[str, int]:
    min_col, min_row, max_col, max_row = _parse_range_for_write(request.range)

    cells_written = 0
    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            column_letter = get_column_letter(col_idx)
            formula = request.formula.replace("{row}", str(row_idx)).replace("{col}", column_letter)
            worksheet.cell(row=row_idx, column=col_idx, value=formula)
            cells_written += 1

    range_applied = f"{worksheet.cell(min_row, min_col).coordinate}:{worksheet.cell(max_row, max_col).coordinate}"
    return range_applied, cells_written


def _compute_aggregate(function: str, values: list[Any]) -> float | int:
    numeric_values: list[float] = []
    for value in values:
        if isinstance(value, str) and value.startswith("="):
            raise InvalidRangeError(
                "Source range contains an uncomputed formula; local evaluation only supports literal values."
            )
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            numeric_values.append(value)

    if function == "SUM":
        return sum(numeric_values)
    if function == "COUNT":
        return len(numeric_values)
    if function == "AVERAGE":
        if not numeric_values:
            raise InvalidRangeError("AVERAGE requires at least one numeric value in the source range.")
        return sum(numeric_values) / len(numeric_values)
    if function == "MIN":
        return min(numeric_values) if numeric_values else 0
    if function == "MAX":
        return max(numeric_values) if numeric_values else 0
    raise ValueError(f"Unsupported function '{function}'.")  # unreachable: validated by the request schema


def _apply_function(worksheet: Worksheet, request: FormulaRequest) -> tuple[str, int, float | int]:
    dest_col, dest_row = _parse_cell_for_write(request.cell)
    src_min_col, src_min_row, src_max_col, src_max_row = _parse_range_for_read(worksheet, request.source_range)

    values = [
        cell.value
        for row in worksheet.iter_rows(
            min_row=src_min_row, max_row=src_max_row, min_col=src_min_col, max_col=src_max_col
        )
        for cell in row
    ]
    computed_value = _compute_aggregate(request.function, values)

    formula = f"={request.function}({request.source_range})"
    worksheet.cell(row=dest_row, column=dest_col, value=formula)

    range_applied = worksheet.cell(dest_row, dest_col).coordinate
    return range_applied, 1, computed_value


def apply_formula(file_id: str, path: Path, request: FormulaRequest) -> tuple[FormulaResult, Workbook]:
    """Load the workbook at `path`, write the requested formula(s) in
    memory, and return both a summary and the mutated Workbook. The
    caller decides whether to persist it (see write_formula_workbook) --
    nothing here touches the file on disk."""
    workbook = openpyxl.load_workbook(path)
    if request.sheet_name not in workbook.sheetnames:
        raise SheetNotFoundError(f"Sheet '{request.sheet_name}' not found in workbook '{file_id}'.")

    worksheet = workbook[request.sheet_name]

    if request.formula is not None:
        range_applied, cells_written = _apply_template(worksheet, request)
        computed_value = None
    else:
        range_applied, cells_written, computed_value = _apply_function(worksheet, request)

    result = FormulaResult(
        file_id=file_id,
        sheet_name=request.sheet_name,
        range_applied=range_applied,
        cells_written=cells_written,
        computed_value=computed_value,
    )
    return result, workbook


def write_formula_workbook(workbook: Workbook) -> bytes:
    """Serialize a (mutated, in-memory) workbook to .xlsx bytes, for committing to storage."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
