"""
Export engine.

Two independent export paths for a stored workbook:
- xlsx: the stored file's raw bytes, served back unchanged -- this is the
  "download the result" step after upload/clean/format/formula/chart, not
  a transformation, so no in-memory workbook load is needed at all.
- csv: a single sheet's data, flattened to CSV text. Read read-only from
  disk (same pattern as workbook/reader.py); values are normalized via
  to_json_safe first, since raw openpyxl cell values (datetime, Decimal,
  ...) aren't directly CSV-writable as text.

Unlike the cleaning/formatting/formula/chart engines, export never
mutates a workbook or writes a new file_id -- it only reads and returns
bytes/text, so there's no preview/commit distinction here.
"""

import csv
import io
from pathlib import Path

import openpyxl

from app.utils.json_safe import to_json_safe


class SheetNotFoundError(Exception):
    """Raised when the requested sheet_name doesn't exist in the workbook."""


def export_xlsx(path: Path) -> bytes:
    """Return the stored workbook's raw .xlsx bytes, unmodified."""
    return path.read_bytes()


def export_sheet_csv(path: Path, sheet_name: str) -> str:
    """Read a single sheet's data (read-only, from disk) and return it as CSV text."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in workbook.")
        worksheet = workbook[sheet_name]

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow([to_json_safe(v) for v in row])
        return buffer.getvalue()
    finally:
        workbook.close()
