"""
Chart engine.

Embeds native Excel chart objects (openpyxl.chart) into a worksheet,
anchored at a given cell and referencing existing cell data by range --
not matplotlib images, since native charts stay editable and linked to
their source data when the user reopens the file in Excel.

Supports five chart types:
- bar, line, area: any number of data series (one column per series, via
  data_range), plus an optional categories_range for axis labels.
- pie: a single data column only (Excel pie charts show one series);
  rejected with InvalidRangeError if data_range spans more than one
  column.
- scatter: unlike the others, needs explicit x_range/y_range rather than
  a single data_range, since openpyxl's ScatterChart series carry
  xvalues and yvalues separately.

Like the cleaning/formatting/formula engines, this only mutates an
in-memory copy of the workbook -- nothing is written to disk unless the
caller commits, and committing always writes a new file rather than
overwriting the original.
"""

import io
from pathlib import Path

import openpyxl
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string, range_boundaries
from openpyxl.utils.exceptions import CellCoordinatesException
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.chart import ChartResult
from app.schemas.chart import ChartRequest

_EXCEL_MAX_ROW = 1_048_576
_EXCEL_MAX_COL = 16_384

# openpyxl doesn't publicly export a common chart base type -- these are
# the only two shapes this module builds, aliased here so signatures
# don't have to spell out the full union.
_SeriesChart = BarChart | LineChart | AreaChart | PieChart
Chart = _SeriesChart | ScatterChart

_SERIES_CHART_CLASSES: dict[str, type[_SeriesChart]] = {
    "bar": BarChart,
    "line": LineChart,
    "area": AreaChart,
    "pie": PieChart,
}


class SheetNotFoundError(Exception):
    """Raised when the requested sheet_name doesn't exist in the workbook."""


class InvalidRangeError(Exception):
    """Raised when a range/cell reference can't be parsed, is out of bounds
    for the sheet's actual data, or is structurally invalid for the
    requested chart type (e.g. a multi-column data_range for a pie
    chart)."""


def _validate_bounds_against_sheet(
    worksheet: Worksheet, min_col: int, min_row: int, max_col: int, max_row: int, ref: str
) -> None:
    sheet_max_row = worksheet.max_row or 1
    sheet_max_col = worksheet.max_column or 1
    if min_row < 1 or min_col < 1 or max_row > sheet_max_row or max_col > sheet_max_col:
        raise InvalidRangeError(
            f"Range '{ref}' is out of bounds for a sheet with {sheet_max_row} row(s) and {sheet_max_col} column(s)."
        )


def _parse_range(worksheet: Worksheet, ref: str) -> tuple[int, int, int, int]:
    try:
        bounds = range_boundaries(ref)
    except ValueError as exc:
        raise InvalidRangeError(f"Invalid range '{ref}': {exc}") from exc
    _validate_bounds_against_sheet(worksheet, *bounds, ref)
    return bounds


def _parse_anchor(ref: str) -> str:
    try:
        column_letter, row = coordinate_from_string(ref)
    except (ValueError, CellCoordinatesException) as exc:
        raise InvalidRangeError(f"Invalid anchor cell '{ref}': {exc}") from exc
    col = column_index_from_string(column_letter)
    if row < 1 or col < 1 or row > _EXCEL_MAX_ROW or col > _EXCEL_MAX_COL:
        raise InvalidRangeError(f"Anchor cell '{ref}' is out of Excel's sheet bounds.")
    return ref


def _build_series_chart(worksheet: Worksheet, request: ChartRequest) -> _SeriesChart:
    assert request.data_range is not None  # enforced by ChartRequest validator
    min_col, min_row, max_col, max_row = _parse_range(worksheet, request.data_range)

    if request.chart_type == "pie" and max_col != min_col:
        raise InvalidRangeError("Pie charts support a single data column; 'data_range' spans more than one column.")

    chart = _SERIES_CHART_CLASSES[request.chart_type]()
    data = Reference(worksheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    chart.add_data(data, titles_from_data=True)

    if request.categories_range is not None:
        cat_min_col, cat_min_row, cat_max_col, cat_max_row = _parse_range(worksheet, request.categories_range)
        categories = Reference(
            worksheet, min_col=cat_min_col, min_row=cat_min_row, max_col=cat_max_col, max_row=cat_max_row
        )
        chart.set_categories(categories)

    return chart


def _build_scatter_chart(worksheet: Worksheet, request: ChartRequest) -> ScatterChart:
    assert request.x_range is not None and request.y_range is not None  # enforced by ChartRequest validator
    x_min_col, x_min_row, x_max_col, x_max_row = _parse_range(worksheet, request.x_range)
    y_min_col, y_min_row, y_max_col, y_max_row = _parse_range(worksheet, request.y_range)

    xvalues = Reference(worksheet, min_col=x_min_col, min_row=x_min_row, max_col=x_max_col, max_row=x_max_row)
    yvalues = Reference(worksheet, min_col=y_min_col, min_row=y_min_row, max_col=y_max_col, max_row=y_max_row)

    chart = ScatterChart()
    chart.series.append(Series(yvalues, xvalues))
    return chart


def build_chart(file_id: str, path: Path, request: ChartRequest) -> tuple[ChartResult, Workbook]:
    """Load the workbook at `path`, build and anchor the requested chart in
    memory, and return both a summary and the mutated Workbook. The caller
    decides whether to persist it (see write_chart_workbook) -- nothing
    here touches the file on disk."""
    workbook = openpyxl.load_workbook(path)
    if request.sheet_name not in workbook.sheetnames:
        raise SheetNotFoundError(f"Sheet '{request.sheet_name}' not found in workbook '{file_id}'.")

    worksheet = workbook[request.sheet_name]
    anchor = _parse_anchor(request.anchor)

    if request.chart_type == "scatter":
        chart = _build_scatter_chart(worksheet, request)
    else:
        chart = _build_series_chart(worksheet, request)

    if request.title is not None:
        chart.title = request.title

    worksheet.add_chart(chart, anchor)

    result = ChartResult(
        file_id=file_id,
        sheet_name=request.sheet_name,
        chart_type=request.chart_type,
        anchor=anchor,
        title=request.title,
    )
    return result, workbook


def write_chart_workbook(workbook: Workbook) -> bytes:
    """Serialize a (mutated, in-memory) workbook to .xlsx bytes, for committing to storage."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
