"""
Data analyzer.

Computes per-column statistics for a workbook: inferred data type, null and
unique counts, and (for numeric columns) min/max/mean/sum. This is a
read-only analysis layer built on top of the same source file the Workbook
Reader reads — it opens the workbook independently and looks at full column
data, since the reader only keeps a preview of rows in memory.
"""

import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.analysis import ColumnAnalysis, SheetAnalysis, WorkbookAnalysis
from app.utils.json_safe import to_json_safe


def _classify(value: Any) -> str:
    """Classify a single non-null cell value into a column data type."""
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float, Decimal)):
        return "numeric"
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return "date"
    return "text"


def _analyze_column(name: Any, index: int, values: list[Any]) -> ColumnAnalysis:
    non_null_values = [v for v in values if v is not None]
    null_count = len(values) - len(non_null_values)

    if not non_null_values:
        return ColumnAnalysis(
            name=name,
            index=index,
            inferred_type="empty",
            null_count=null_count,
            unique_count=0,
        )

    types = {_classify(v) for v in non_null_values}
    inferred_type = types.pop() if len(types) == 1 else "mixed"
    unique_count = len(set(non_null_values))

    min_value = max_value = mean_value = sum_value = None
    if inferred_type == "numeric":
        numeric_values = [float(v) for v in non_null_values]
        min_value = to_json_safe(min(non_null_values))
        max_value = to_json_safe(max(non_null_values))
        sum_value = sum(numeric_values)
        mean_value = sum_value / len(numeric_values)

    return ColumnAnalysis(
        name=name,
        index=index,
        inferred_type=inferred_type,
        null_count=null_count,
        unique_count=unique_count,
        min=min_value,
        max=max_value,
        mean=mean_value,
        sum=sum_value,
    )


def _analyze_sheet(worksheet: Worksheet) -> SheetAnalysis:
    rows_iter = worksheet.iter_rows(values_only=True)

    headers: list[Any] = []
    first_row = next(rows_iter, None)
    if first_row is not None:
        headers = [to_json_safe(v) for v in first_row]

    data_rows = list(rows_iter)
    column_count = worksheet.max_column or 0

    columns = [
        _analyze_column(
            name=headers[index] if index < len(headers) else None,
            index=index,
            values=[row[index] if index < len(row) else None for row in data_rows],
        )
        for index in range(column_count)
    ]

    return SheetAnalysis(name=worksheet.title, columns=columns)


def analyze_workbook(file_id: str, path: Path) -> WorkbookAnalysis:
    """Load the workbook at `path` and return per-column analysis for each sheet."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        sheets = [_analyze_sheet(workbook[name]) for name in sheet_names]
    finally:
        workbook.close()

    return WorkbookAnalysis(file_id=file_id, sheet_names=sheet_names, sheets=sheets)
