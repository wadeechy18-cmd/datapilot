"""
Workbook reader.

Loads an .xlsx file from disk into the internal WorkbookInfo
representation (models/workbook.py). This is read-only: it describes the
structure and a preview of the data. It does not analyze, clean, or
modify anything — those are separate, later engines (Data Analyzer,
Cleaning Engine, ...) that will each take a workbook and act on it.
"""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.workbook import SheetInfo, WorkbookInfo
from app.utils.json_safe import to_json_safe

# No longer used to cap the main reader's preview_rows (the grid shows every
# row now) -- kept as the sample size for the Cleaning engine's own
# pending-change preview response, a separate, smaller concern.
PREVIEW_ROW_COUNT = 10


def _read_sheet(worksheet: Worksheet) -> SheetInfo:
    row_count = worksheet.max_row or 0
    column_count = worksheet.max_column or 0

    rows_iter = worksheet.iter_rows(values_only=True)

    headers: list[Any] = []
    first_row = next(rows_iter, None)
    if first_row is not None:
        headers = [to_json_safe(v) for v in first_row]

    preview_rows: list[list[Any]] = []
    non_empty_cells = 0
    empty_cells = 0
    numeric_cells = 0
    text_cells = 0

    for row in rows_iter:
        preview_rows.append([to_json_safe(v) for v in row])

        for value in row:
            if value is None:
                empty_cells += 1
            else:
                non_empty_cells += 1
                if isinstance(value, (int, float, bool)):
                    numeric_cells += 1
                elif isinstance(value, str):
                    text_cells += 1
                else:
                    text_cells += 1

    return SheetInfo(
        name=worksheet.title,
        row_count=row_count,
        column_count=column_count,
        headers=headers,
        preview_rows=preview_rows,
        non_empty_cells=non_empty_cells,
        empty_cells=empty_cells,
        numeric_cells=numeric_cells,
        text_cells=text_cells,
    )


def read_workbook(file_id: str, path: Path) -> WorkbookInfo:
    """Load the workbook at `path` and return its structural summary."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        sheets = [_read_sheet(workbook[name]) for name in sheet_names]
    finally:
        workbook.close()

    return WorkbookInfo(
        file_id=file_id,
        sheet_names=sheet_names,
        sheets=sheets,
    )
